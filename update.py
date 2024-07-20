import openpyxl
import fnmatch
import os
GREEN = '\033[92m'  # 绿色
RED = '\033[91m'  # 红色
RESET = '\033[0m'    # 重置颜色

file_paths = []
for file in os.listdir('.'):
    if fnmatch.fnmatch(file, '*计算机*.xlsx'):
        file_paths.append(file)
try:
    try:
        # 遍历匹配到的文件
        for file_path in file_paths:
            if "总表" in file_path:
                workbook = openpyxl.load_workbook(file_path)
                current_path=file_path
                #print(file_path)
        sheet = workbook.active
    except Exception as e:
        print(f"{RED}[-]文件名不匹配或不存在！{RESET}")
        exit(-1)
    print("[+]更新总表数据")

    # print(sheet.max_row)

    # 优化后的计算总分的函数
    def calculate_total(row_data):
        total = sum(float(cell.value) for cell in row_data[3:-1] if cell.value is not None)  # 第4列到倒数第二列的数据求和
        return total

    # 批量写入总分
    updated_rows = 0  # 记录更新的行数
    for row in range(4, sheet.max_row - 9):  # 从第四行到倒数第九行的数据
        row_data = list(sheet[row])  # 读取整行数据
        current_total = float(row_data[-1].value) if row_data[-1].value is not None else 0  # 获取当前总分
        new_total = calculate_total(row_data)
        if new_total != current_total:  # 检查是否需要更新
            sheet.cell(row=row, column=22, value=new_total)  # 写入总分
            updated_rows += 1

    # 保存修改后的Excel文件
    workbook.save(current_path)
    print(f"{GREEN}[+]更新总表数据完成，已更新{updated_rows}条数据！{RESET}")

except Exception as e:
    print(f"{RED}[-]update.py->出错！{RESET}")
    exit(-1)