# -*- coding: utf-8 -*-
import openpyxl
import threading
import os
import sys
import fnmatch
from openpyxl.styles import Border, Side

GREEN = '\033[92m'
RED = '\033[91m'
RESET = '\033[0m'

class Excel_Process:
    def __init__(self, data_file='./data.txt'):
        self.data_file = data_file
        self.file_paths = []
        self.flag_cnt = 0
        self.current_week = 0
        self.class_list = set()
        self.absent_list = []
        self.late_list = []
        self.total_list = []
        self.week_table_students = []
        self.absence_summary = {}
        self.current_class = 2233
        self.total_list_max_row = 0
        self.week_day_mapping = {'周一': 1, '周二': 2, '周三': 3, '周四': 4, '周五': 5, '周六': 6, '周日': 7}
        self.total_workbook = None
        self.week_workbook = None
        self.analytics_workbook = None
        self.analytics_sheet = None

    def print_banner(self):
        pattern = """////////////////////////////////////////////////////////////////////
//                          _ooOoo_                               //
//                         o8888888o                              //
//                         88" . "88                              //
//                         (| ^_^ |)                              //
//                         O\  =  /O                              //
//                      ____/`---'\____                           //
//                    .'  \\|     |//  `.                          //
//                   /  \\|||  :  |||//  \                         //
//                  /  _||||| -:- |||||-  \                       //
//                  |   | \\\  -  /// |   |                        //
//                  | \_|  ''\---/''  |   |                       //
//                  \  .-\__  `-`  ___/-. /                       //
//                ___`. .'  /--.--\  `. . ___                     //
//              ."" '<  `.___\_<|>_/___.'  >'"".                  //
//            | | :  `- \`.;`\ _ /`;.`/ - ` : | |                 //
//            \  \ `-.   \_ __\ /__ _/   .-` /  /                 //
//      ========`-.____`-.___\_____/___.-`____.-'========         //
//                           `=---='                              //
//      ^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^        //
//             佛祖保佑       永不宕机      永无BUG               //
//                                                                //
//@Author null-sky                                                //
//@Date 2024/8/12                                                 //
//@GitHub https://github.com/c0yt                                 //
//@version 8.0                                                    //
//@Name Auto_Excel                                                //
////////////////////////////////////////////////////////////////////
"""
        print(f"{GREEN}{pattern}{RESET}")

    def convert_file_encoding(self):
        try:
            if not os.path.exists(self.data_file):
                print(f"{RED}[-]data.txt不存在！")
                sys.exit()
            with open(self.data_file, 'r', encoding='ansi') as f:
                content = f.read()
            utf8_content = content.encode('utf-8')
            with open(self.data_file, 'w', encoding='utf-8') as f:
                f.write(utf8_content.decode('utf-8'))
            print(f"{GREEN}[+]文件已转换为utf-8编码！{RESET}")
        except Exception as e:
            print(f"{GREEN}[+]文件已经是为utf-8编码哦~{RESET}")

    def load_files(self):
        for file in os.listdir('.'):
            if fnmatch.fnmatch(file, '*计算机*.xlsx'):
                self.file_paths.append(file)
        self._load_workbooks()

    def _load_workbooks(self):
        try:
            for file_path in self.file_paths:
                if "总表" in file_path:
                    self.total_workbook = openpyxl.load_workbook(file_path, data_only=False)
                    self.current_total_path = file_path
                    print(f'{GREEN}[+]当前载入总表：{self.current_total_path}{RESET}')
                    self.flag_cnt += 1
                elif "周表" in file_path:
                    self.week_workbook = openpyxl.load_workbook(file_path, data_only=True)
                    self.current_week_path = file_path
                    print(f'{GREEN}[+]当前载入周表：{self.current_week_path}{RESET}')
                    self.flag_cnt += 1
                elif "考勤统计表" in file_path:
                    self.analytics_workbook = openpyxl.load_workbook(file_path,data_only=False)
                    self.analytics_path = file_path
                    print(f'{GREEN}[+]当前载入统计表：{self.analytics_path}{RESET}')
                    self.flag_cnt += 1

            if self.flag_cnt != 3:
                print(f"{RED}[-]当前目录存在多个重名文件或文件缺失！{RESET}")
                sys.exit()

            self.analytics_sheet = self.analytics_workbook[self.analytics_workbook.sheetnames[0]]
            self.total_sheet = self.total_workbook[self.total_workbook.sheetnames[0]]
            self.week_sheet = self.week_workbook[self.week_workbook.sheetnames[0]]
            self.total_list_max_row = self.total_sheet.max_row - 12

            print("[+]获取班级列表")
            for row in self.total_sheet.iter_rows(min_row=4, max_row=self.total_list_max_row, min_col=1, max_col=1):
                if str(row[0].value).isdigit():
                    self.class_list.add(row[0].value)

            self.class_list = list(self.class_list)

            with open('class.txt', 'w', encoding='utf-8') as f:
                for cls in sorted(self.class_list):
                    f.write(f"{cls}\n")
            print(f"{GREEN}[+]class.txt更新成功！{RESET}")

        except Exception as e:
            print(f"{RED}[-]文件名不匹配或不存在！请关闭文件后重试！{RESET}")
            sys.exit()

    class MyThread(threading.Thread):
        def __init__(self, func, args=()):
            super().__init__()
            self.func = func
            self.args = args

        def run(self):
            self.result = self.func(*self.args)

        def get_result(self):
            threading.Thread.join(self)
            try:
                return self.result
            except Exception:
                return None

    def find_student_id(self, class_id, name):
        matched_students = []
        for student in self.total_list:
            if student[0] == class_id:
                if student[2] == name:
                    matched_students.append(student)
            elif matched_students:
                break
        if len(matched_students) == 0:
            print(f"{RED}[-]未找到：{class_id}班{name}{RESET}")
            sys.exit()
        elif len(matched_students) == 1:
            return matched_students[0][1]
        else:
            print(f"{RED}[!]{class_id}班存在多个同名学生：{name}{RESET}")
            print("+" + "-" * 30 + "+")
            print(f"| {'序号':<3} | {'学号':<9} | {'姓名':<6} |")
            print("+" + "-" * 30 + "+")

            for i, student in enumerate(matched_students, 1):
                print(f"| {i:<4} | {student[1]:<10} | {student[2]:<6} |")

            print("+" + "-" * 30 + "+")

            while True:
                try:
                    choice = int(input(f"[!]请选择正确的学号（输入1-{len(matched_students)}的数字）："))
                    if 1 <= choice <= len(matched_students):
                        return matched_students[choice - 1][1]
                    else:
                        print(f"{RED}[-]输入不合法，请输入1-{len(matched_students)}之间的数字{RESET}")
                except ValueError:
                    print(f"{RED}[-]输入不合法，请输入一个数字{RESET}")

    def add_to_list(self, record, flag):
        records = record.split(',')
        class_id = None
        for item in records:
            parts = item.split(":")
            if parts[0][0].isdigit():
                class_id = int(parts[0][0:4])
                thread = self.MyThread(self.find_student_id, (class_id, parts[1]))
                thread.start()
                thread.join()
                if flag == 1:
                    self.absent_list.append(int(parts[0][0:4]))
                    self.absent_list.append([thread.get_result()] + parts[1:])
                else:
                    self.late_list.append(int(parts[0][0:4]))
                    self.late_list.append([thread.get_result()] + parts[1:])
            else:
                thread = self.MyThread(self.find_student_id, (class_id, parts[0]))
                thread.start()
                thread.join()
                if flag == 1:
                    self.absent_list.append([thread.get_result()] + parts)
                else:
                    self.late_list.append([thread.get_result()] + parts)

    def write_week_table(self, data_list, value):
        start_row = 3
        start_column = 3
        num = 1
        for record in data_list:
            total = 0
            if record in self.class_list:
                self.current_class = record
                continue
            row = num
            if record[0] in self.week_table_students:
                row = self.week_table_students.index(record[0]) + 1
            else:
                self.week_table_students.append(record[0])
                row = len(self.week_table_students)
                current_row = start_row + len(self.week_table_students)
                self.week_sheet.insert_rows(idx=current_row, amount=1)
                self.week_sheet['A' + str(current_row)].value = int(self.current_class)
                self.week_sheet['B' + str(current_row)].value = int(record[0])
                self.week_sheet['C' + str(current_row)].value = str(record[1])
            current_row = start_row + row
            days = record[2].split(' ')
            day_offset = 0
            for day in days:
                if day.isdigit():
                    column = openpyxl.utils.get_column_letter(day_offset + int(day) + start_column)
                    cell = column + str(current_row)
                    self.week_sheet[cell].value = value
                else:
                    self.week_sheet['BV' + str(current_row)].value = '=SUM(D{0}:BU{0})'.format(current_row)
                    day_offset = (self.week_day_mapping[day] - 1) * 10
            for cell in self.week_sheet.iter_rows(min_row=current_row, max_row=current_row, min_col=start_column + 1,
                                                 max_col=openpyxl.utils.column_index_from_string('BU')):
                for j in cell:
                    if j.value is not None:
                        total += float(j.value)
            self.absence_summary[record[0]] = total
            num += 1

    def write_total_table(self, thread_num, column):
        num = 0
        for student_id, total_absence in self.absence_summary.items():
            if num >= 10 * thread_num and num < (thread_num + 1) * 10:
                row_num = 1
                for student in self.total_list:
                    if student_id in student:
                        self.total_sheet.cell(row_num + 3, column).value = total_absence
                    row_num += 1
            num += 1

    def calculate_total(self, row_data):
        total = sum(float(cell.value) for cell in row_data[3:-1] if cell.value is not None)
        return total

    def process_data_file(self):
        try:
            print("[+]获取总名单")
            for row in self.total_sheet.iter_rows(min_row=4, max_row=self.total_list_max_row, min_col=1, max_col=3):
                if str(row[0].value).isdigit():
                    self.total_list.append([row[0].value, row[1].value, row[2].value])

            try:
                with open(self.data_file, 'r', encoding='utf-8') as f:
                    flag = 0
                    for line in f:
                        stripped_line = line.strip()
                        if "周数" in stripped_line:
                            self.current_week = int(stripped_line[3:])
                            print("[+]当前周数:", self.current_week)
                            continue
                        if "旷课" in stripped_line:
                            flag = 1
                            continue
                        if "迟到" in stripped_line:
                            flag = 2
                            continue
                        if "请假" in stripped_line:
                            break
                        self.add_to_list(stripped_line, flag)
            except Exception as e:
                print(e)
                raise

            try:
                print("[+]写入周表")
                week_max_row = self.week_sheet.max_row - 16
                if week_max_row != 4:
                    while week_max_row > 4:
                        self.week_sheet.delete_rows(4)
                        week_max_row -= 1

                    print(f"{GREEN}[+]周表重置成功！{RESET}")
                else:
                    pass

                self.write_week_table(self.absent_list, 1)
                self.write_week_table(self.late_list, 0.25)

                border_style = Border(
                    left=Side(border_style='thin'),
                    right=Side(border_style='thin'),
                    top=Side(border_style='thin'),
                    bottom=Side(border_style='thin')
                )

                start_row = 4
                end_row = self.week_sheet.max_row - 17
                start_col = 1
                end_col = self.week_sheet.max_column -9
                for row in self.week_sheet.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
                    for cell in row:
                        cell.border = border_style


                print(f"{GREEN}[+]周表写入成功！{RESET}")
            except Exception as e:
                print(f"{RED}[-]周表写入失败！{RESET}")
                print(f"{RED}[-]请检查data.txt")
                sys.exit()

            start_row = 3
            start_column = 3
            week_column = int(self.current_week) + start_row


            print("[+]写入总表")
            try:
                threads = []
                for i in range(0, int(len(self.absence_summary) / 10) + 1):
                    thread = threading.Thread(target=self.write_total_table, args=(i, week_column))
                    threads.append(thread)

                for thread in threads:
                    thread.start()

                for thread in threads:
                    thread.join()

                print(f"{GREEN}[+]总表写入成功！{RESET}")
            except Exception as e:
                print(f"{RED}[-]总表写入失败！{RESET}")
                print(f"{RED}[-]请检查data.txt")
                sys.exit()

            updated_rows = 0
            print("[+]更新总表数据")
            for row in range(4, self.total_list_max_row):
                row_data = list(self.total_sheet[row])
                current_total = float(row_data[-1].value) if row_data[-1].value is not None else 0
                new_total = self.calculate_total(row_data)
                if new_total != current_total:
                    self.total_sheet.cell(row=row, column=22, value=new_total)
                    updated_rows += 1

            print(f"{GREEN}[+]更新总表数据完成，已更新{updated_rows}条数据！{RESET}")

            self._write_analytics_sheet()
            self.week_workbook.save(self.current_week_path)
            self.total_workbook.save(self.current_total_path)

        except Exception as e:
            print(f"{RED}[-]process_data_file->出错!{RESET}")
            sys.exit()

    def _write_analytics_sheet(self):
        class_dict = {}

        def add_list(line, flag):
            str1 = line.split(',')
            current_class = None
            for r in str1:
                rr = r.split(':')
                if rr[0][0].isdigit():
                    current_class = int(rr[0][0:4])
                    name = rr[1]
                    if current_class not in class_dict:
                        class_dict[current_class] = {'旷课': [], '迟到': [], '请假': []}
                else:
                    name = rr[0]
                if flag == 1:
                    class_dict[current_class]['旷课'].append(name)
                elif flag == 2:
                    class_dict[current_class]['迟到'].append(name)
                elif flag == 3:
                    class_dict[current_class]['请假'].append(name)

        with open(self.data_file, 'r', encoding='utf-8') as file:
            flag = 0
            for line in file:
                newline = line.strip()
                if "旷课" in line:
                    flag = 1
                    continue
                if "迟到" in line:
                    flag = 2
                    continue
                if "请假" in line:
                    flag = 3
                    continue
                add_list(newline, flag)

        sorted_class_dict = dict(sorted(class_dict.items(), reverse=True))

        absent_str = ""
        for class_id, attendance_data in sorted_class_dict.items():
            if attendance_data['旷课']:
                absent_str += f'{class_id}：'
                absent_str += "、".join(attendance_data['旷课'])
                absent_str += "\n"

        late_str = ""
        for class_id, attendance_data in sorted_class_dict.items():
            if attendance_data['迟到']:
                late_str += f'{class_id}：'
                late_str += "、".join(attendance_data['迟到'])
                late_str += "\n"

        leave_str_d = ""
        for class_id, attendance_data in sorted_class_dict.items():
            if attendance_data['请假']:
                leave_str_d += f'{class_id}：'
                leave_str_d += "、".join(attendance_data['请假'])
                leave_str_d += "\n"

        all_str = "旷课：" + "\n" + absent_str + "\n" + "迟到：" + "\n" + late_str
        leave_str = "请假：" + "\n" + leave_str_d

        print("[+]写入统计表")
        self.analytics_sheet['B9'] = all_str
        self.analytics_sheet['M9'] = leave_str

        self.analytics_workbook.save(self.analytics_path)
        print(f"{GREEN}[+]All success!{RESET}")


if __name__ == "__main__":
    # 实例化对象
    processor = Excel_Process()
    # 展示声明
    processor.print_banner()
    # 修正文件编码
    processor.convert_file_encoding()
    # 载入表格
    processor.load_files()
    # 处理数据
    processor.process_data_file()
